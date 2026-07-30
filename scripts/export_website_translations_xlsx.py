"""Export all website locale content to a review-friendly Excel workbook."""

from __future__ import annotations

import json
import ast
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
LOCALES_DIR = ROOT / "Frontend" / "public" / "locales"
OUTPUT = ROOT / "website-translations-all-content.xlsx"
LANGUAGES = [
    ("English", "en.json"),
    ("Khmer", "kh.json"),
    ("Chinese", "zh.json"),
    ("Korean", "ko.json"),
]
ABOUT_SOURCE = ROOT / "Frontend" / "src" / "pages" / "About" / "index.tsx"
TERMS_SOURCE = ROOT / "Frontend" / "src" / "pages" / "Terms" / "index.tsx"


def flatten(value: Any, path: tuple[str, ...] = ()) -> dict[str, Any]:
    rows: dict[str, Any] = {}
    if isinstance(value, dict):
        for key, child in value.items():
            rows.update(flatten(child, (*path, str(key))))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            rows.update(flatten(child, (*path, str(index))))
    else:
        rows[".".join(path)] = value
    return rows


def display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def string_literals(source: str) -> list[str]:
    pattern = re.compile(r"'(?:\\.|[^'\\])*'")
    return [ast.literal_eval(match.group(0)) for match in pattern.finditer(source)]


def inline_source_content() -> tuple[dict[str, str], dict[str, Any], dict[str, Any]]:
    about_source = ABOUT_SOURCE.read_text(encoding="utf-8")
    about_block = about_source.split("const khmerCopy", 1)[1].split("\n};", 1)[0]
    pairs = re.findall(
        r"('(?:\\.|[^'\\])*')\s*:\s*('(?:\\.|[^'\\])*')",
        about_block,
    )
    about_khmer = {ast.literal_eval(key): ast.literal_eval(value) for key, value in pairs}

    terms_source = TERMS_SOURCE.read_text(encoding="utf-8")
    english_block = terms_source.split("const englishSections", 1)[1].split("const khmerSections", 1)[0]
    khmer_block = terms_source.split("const khmerSections", 1)[1].split("export default", 1)[0]
    english_values = string_literals(english_block)
    khmer_values = string_literals(khmer_block)
    english_terms = {
        "eyebrow": "Important Booking Information",
        "title": "Terms and Conditions",
        "intro": "Please review the following terms before confirming your booking.",
        "sections": [
            {"title": english_values[index], "paragraphs": [english_values[index + 1]]}
            for index in range(0, len(english_values), 2)
        ],
    }
    khmer_headers = [
        ast.literal_eval(value)
        for value in re.findall(r"isKhmer\s*\?\s*('(?:\\.|[^'\\])*')", terms_source)
    ]
    khmer_terms = {
        "eyebrow": khmer_headers[0],
        "title": khmer_headers[1],
        "intro": khmer_headers[2],
        "sections": [
            {"title": khmer_values[index], "paragraphs": [khmer_values[index + 1]]}
            for index in range(0, len(khmer_values), 2)
        ],
    }
    return about_khmer, english_terms, khmer_terms


def main() -> None:
    dictionaries = {
        language: flatten(json.loads((LOCALES_DIR / filename).read_text(encoding="utf-8")))
        for language, filename in LANGUAGES
    }
    about_khmer, english_terms, khmer_terms = inline_source_content()
    chinese_about = json.loads((LOCALES_DIR / "zh.json").read_text(encoding="utf-8")).get("aboutInline", {})
    for english_text in chinese_about:
        dictionaries["English"][f"aboutInline.{english_text}"] = english_text
        dictionaries["Khmer"][f"aboutInline.{english_text}"] = about_khmer.get(english_text, english_text)
    dictionaries["English"].update(
        {f"termsInline.{key}": value for key, value in flatten(english_terms).items()}
    )
    dictionaries["Khmer"].update(
        {f"termsInline.{key}": value for key, value in flatten(khmer_terms).items()}
    )
    keys = sorted(set().union(*(dictionary.keys() for dictionary in dictionaries.values())))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All Translations"
    headers = [*(language for language, _ in LANGUAGES), "Translation Key"]
    sheet.append(headers)

    for key in keys:
        sheet.append([*(display(dictionaries[language].get(key)) for language, _ in LANGUAGES), key])

    header_fill = PatternFill("solid", fgColor="6B9158")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 25
    widths = [55, 55, 55, 55, 52]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = workbook.create_sheet("Coverage Summary")
    summary.append(["Language", "Translated values", "Total keys", "Coverage"])
    for language, _ in LANGUAGES:
        count = sum(1 for key in keys if display(dictionaries[language].get(key)).strip())
        summary.append([language, count, len(keys), count / len(keys) if keys else 0])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in summary["D"][1:]:
        cell.number_format = "0.00%"
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 15
    summary.column_dimensions["D"].width = 15

    workbook.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(keys)} translation keys")


if __name__ == "__main__":
    main()
